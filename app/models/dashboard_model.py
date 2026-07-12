from datetime import datetime
from time import sleep
import os
import io
import tempfile

from PyQt6.QtCore import QObject, pyqtSignal, QThreadPool
from PyQt6.QtWidgets import QMessageBox

from app.models.database.defect_dao import DefectDAO
from app.models.database.inference_dao import InferenceDAO
from app.models.utils import Worker, WorkerSignals
from openpyxl.utils import get_column_letter


class DashboardModel(QObject):
    data_updated = pyqtSignal(dict)
    excel_export_finished = pyqtSignal(str)  # 导出完成信号，传递文件路径
    excel_export_error = pyqtSignal(str)     # 导出错误信号，传递错误信息

    def __init__(self):
        super().__init__()
        self.inference_dao = InferenceDAO()
        self.defect_dao = DefectDAO()
        self.thread_pool = QThreadPool() # 初始化线程池
        self.thread_pool.setMaxThreadCount(2)


    def get_monthly_stats(self, year, month):
        # 使用多线程异步获取数据，避免阻塞
        worker = Worker(self._get_monthly_stats_task, year, month)
        worker.signals.result.connect(self.data_updated.emit)
        self.thread_pool.start(worker)    
    def _get_monthly_stats_task(self, year, month, signals: WorkerSignals):
        # 将原有同步查询逻辑迁移到此处
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        total, results = self.inference_dao.fetch_inference_results_by_page(
            page=1,
            page_size=999999999,
            from_time=start_date.strftime("%Y-%m-%d %H:%M:%S"),
            to_time=end_date.strftime("%Y-%m-%d %H:%M:%S")
        )

        daily_counts = {}
        defect_types = {}
        defective_inspections = 0  # 检测出缺陷的检测次数
        
        # 处理查询结果
        for result in results:
            try:
                # 解析时间戳，提取日期
                cur_date = datetime.strptime(result['timestamp'], "%Y-%m-%d %H:%M:%S").date()
                daily_counts[cur_date] = daily_counts.get(cur_date, 0) + result['defect_count']
                
                # 如果该次检测发现了缺陷，则计入defective_inspections
                if result['defect_count'] > 0:
                    defective_inspections += 1
                
                # 获取该推理结果的缺陷类型统计
                defects = self.defect_dao.fetch_defect_counts_by_inference_id(result['id'])
                for defect in defects:
                    defect_name = defect['defect_name']
                    defect_types[defect_name] = defect_types.get(defect_name, 0) + defect['count']
            except Exception as e:
                print(f"处理推理结果时出错: {e}")
        return {
            'daily_counts': daily_counts,
            'defect_types': defect_types,
            'total_inspections': total,
            'defective_inspections': defective_inspections  # 添加检测出缺陷的检测次数
        }
    
    def export_monthly_data_to_excel(self, year, month):
        """导出指定年月的详细数据到Excel文件"""
        worker = Worker(self._export_excel_task, year, month)
        worker.signals.result.connect(self.excel_export_finished.emit)
        worker.signals.error.connect(self.excel_export_error.emit)
        self.thread_pool.start(worker)    
    def _export_excel_task(self, year, month, signals: WorkerSignals):
        """Excel导出任务的具体实现"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as XLImage
            
            # 获取当月数据
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month + 1, 1)
                
            total, results = self.inference_dao.fetch_inference_results_by_page(
                page=1,
                page_size=999999999,
                from_time=start_date.strftime("%Y-%m-%d %H:%M:%S"),
                to_time=end_date.strftime("%Y-%m-%d %H:%M:%S")
            )
              
            # 创建工作簿和工作表
            wb = openpyxl.Workbook()
            
            # 删除默认工作表
            if wb.active:
                wb.remove(wb.active)
            
            # 创建汇总工作表
            summary_ws = wb.create_sheet("数据汇总")
              # 创建详细数据工作表（包含图片）
            detail_ws = wb.create_sheet("详细检测记录")
            
            # 创建缺陷统计工作表
            defect_ws = wb.create_sheet("缺陷统计")
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 填充汇总工作表
            self._fill_summary_sheet(summary_ws, year, month, total, results, 
                                   header_font, header_fill, header_alignment, border)
              # 填充详细数据工作表（包含图片）
            temp_files = self._fill_detail_sheet_with_images(detail_ws, results, 
                                                           header_font, header_fill, header_alignment, border)
            
            # 填充缺陷统计工作表
            self._fill_defect_sheet(defect_ws, results, 
                                  header_font, header_fill, header_alignment, border)
              
            # 保存文件
            user_home = os.path.expanduser("~")
            export_dir = os.path.join(user_home, "SteelDefect", "exports")
            os.makedirs(export_dir, exist_ok=True)
            
            current_time = datetime.now().strftime("%Y%m%d%H%M%S")
            filename = f"钢材缺陷检测数据_{year}年{month:02d}月_{current_time}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            wb.save(filepath)
            
            # 在Excel保存完成后清理临时文件
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except Exception as e:
                    print(f"清理临时文件失败: {e}")
            
            return filepath
            
        except ImportError:
            raise Exception("缺少openpyxl库，请确保已安装: pip install openpyxl")
        except Exception as e:
            raise Exception(f"导出Excel文件时发生错误: {str(e)}")
    def _fill_summary_sheet(self, ws, year, month, total, results, header_font, header_fill, header_alignment, border):
        """填充汇总工作表"""
        from openpyxl.styles import Font, Alignment
        
        # 标题
        ws['A1'] = f"{year}年{month:02d}月钢材缺陷检测数据汇总"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # 基本统计信息
        ws['A3'] = "统计项目"
        ws['B3'] = "数值"
        for cell in ['A3', 'B3']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].alignment = header_alignment
            ws[cell].border = border
        
        ws['A4'] = "总检测次数"
        ws['B4'] = total
        ws['A5'] = "检测时间范围"
        ws['B5'] = f"{year}-{month:02d}-01 至 {year}-{month:02d}-{(datetime(year, month+1, 1) - datetime(year, month, 1)).days if month < 12 else 31}"
          # 计算缺陷统计
        total_defects = sum(result['defect_count'] for result in results)
        defective_inspections = sum(1 for result in results if result['defect_count'] > 0)
        defect_rate = (defective_inspections / total * 100) if total > 0 else 0
        
        ws['A6'] = "总缺陷数量"
        ws['B6'] = total_defects
        ws['A7'] = "缺陷检出率"
        ws['B7'] = f"{defect_rate:.2f}%"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
    def _fill_detail_sheet_with_images(self, ws, results, header_font, header_fill, header_alignment, border):
        """填充详细数据工作表（包含图片）"""
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, Alignment
        
        # 表头
        headers = ['序号', '检测ID', '检测时间', '原始图片', '标注图片', '缺陷数量', '具体缺陷类型']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8   # 序号
        ws.column_dimensions['B'].width = 12  # 检测ID
        ws.column_dimensions['C'].width = 20  # 检测时间
        ws.column_dimensions['D'].width = 25  # 原始图片
        ws.column_dimensions['E'].width = 25  # 标注图片
        ws.column_dimensions['F'].width = 12  # 缺陷数量
        ws.column_dimensions['G'].width = 40  # 具体缺陷类型
        
        # 设置行高以容纳图片
        row_height = 150
        temp_files = []  # 用于跟踪临时文件
        
        # 填充数据
        for row_idx, result in enumerate(results, 2):
            # 设置行高
            ws.row_dimensions[row_idx].height = row_height
            
            # 基本信息
            ws.cell(row=row_idx, column=1, value=row_idx-1)  # 序号
            ws.cell(row=row_idx, column=2, value=result['id'])  # 检测ID
            ws.cell(row=row_idx, column=3, value=result['timestamp'])  # 检测时间
            ws.cell(row=row_idx, column=6, value=result['defect_count'])  # 缺陷数量
            
            # 获取具体缺陷类型
            defects = self.defect_dao.fetch_defect_counts_by_inference_id(result['id'])
            defect_details = []
            for defect in defects:
                defect_details.append(f"{defect['defect_name']}({defect['count']}个)")
            ws.cell(row=row_idx, column=7, value="; ".join(defect_details))
            
            # 获取完整的推理结果（包含图片数据）
            full_result = self.inference_dao.get_inference_result_by_id(result['id'])
            if full_result:
                # 处理原始图片
                if full_result.get('original_image'):
                    try:
                        original_temp = self._create_temp_image(full_result['original_image'])
                        if original_temp:
                            temp_files.append(original_temp)
                            original_img = XLImage(original_temp)
                            # 调整图片大小以适应单元格
                            original_img.width = 180  # 像素
                            original_img.height = 130  # 像素
                            # 将图片锚定到单元格
                            original_img.anchor = f'D{row_idx}'
                            ws.add_image(original_img)
                    except Exception as e:
                        print(f"添加原始图片失败 (ID: {result['id']}): {e}")
                        ws.cell(row=row_idx, column=4, value="图片加载失败")
                else:
                    ws.cell(row=row_idx, column=4, value="无原始图片")
                
                # 处理标注图片
                if full_result.get('annotated_image'):
                    try:
                        annotated_temp = self._create_temp_image(full_result['annotated_image'])
                        if annotated_temp:
                            temp_files.append(annotated_temp)
                            annotated_img = XLImage(annotated_temp)
                            # 调整图片大小以适应单元格
                            annotated_img.width = 180  # 像素
                            annotated_img.height = 130  # 像素
                            # 将图片锚定到单元格
                            annotated_img.anchor = f'E{row_idx}'
                            ws.add_image(annotated_img)
                    except Exception as e:
                        print(f"添加标注图片失败 (ID: {result['id']}): {e}")
                        ws.cell(row=row_idx, column=5, value="图片加载失败")
                else:
                    ws.cell(row=row_idx, column=5, value="无标注图片")
            else:
                ws.cell(row=row_idx, column=4, value="无法获取图片数据")
                ws.cell(row=row_idx, column=5, value="无法获取图片数据")
            
            # 设置边框和对齐
            for col in range(1, 8):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 返回临时文件列表，在Excel保存后清理
        return temp_files
    def _fill_defect_sheet(self, ws, results, header_font, header_fill, header_alignment, border):
        """填充缺陷统计工作表"""
        from openpyxl.styles import Font
        
        # 统计所有缺陷类型
        defect_summary = {}
        for result in results:
            defects = self.defect_dao.fetch_defect_counts_by_inference_id(result['id'])
            for defect in defects:
                defect_name = defect['defect_name']
                if defect_name not in defect_summary:
                    defect_summary[defect_name] = 0
                defect_summary[defect_name] += defect['count']
        
        # 表头
        headers = ['缺陷类型', '总数量', '占比(%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 填充数据
        total_defects = sum(defect_summary.values())
        sorted_defects = sorted(defect_summary.items(), key=lambda x: x[1], reverse=True)
        
        for row_idx, (defect_name, count) in enumerate(sorted_defects, 2):
            percentage = (count / total_defects * 100) if total_defects > 0 else 0
            
            ws.cell(row=row_idx, column=1, value=defect_name)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=f"{percentage:.2f}%")
            
            # 设置边框
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).border = border
        
        # 添加合计行
        if sorted_defects:
            total_row = len(sorted_defects) + 2
            ws.cell(row=total_row, column=1, value="合计")
            ws.cell(row=total_row, column=2, value=total_defects)
            ws.cell(row=total_row, column=3, value="100.00%")
            
            # 设置合计行样式
            for col in range(1, 4):
                cell = ws.cell(row=total_row, column=col)
                cell.font = Font(bold=True)
                cell.border = border        # 设置列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_temp_image(self, image_blob):
        """从BLOB数据创建临时图片文件"""
        try:
            if not image_blob:
                return None
            
            # 创建临时文件
            with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as temp_file:
                temp_file.write(image_blob)
                temp_path = temp_file.name
            
            return temp_path
        except Exception as e:
            print(f"创建临时图片文件失败: {e}")
            return None